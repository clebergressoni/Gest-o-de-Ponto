import customtkinter as ctk
from tkinter import messagebox
from openpyxl import Workbook
import openpyxl
import pathlib
from datetime import datetime, timedelta

ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

FUNCIONARIOS = ['FUNCIONARIO 1', 'FUNCIONARIO 2']
SENHA = 'Senha_123'

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.layout_config()
        self.sistema()

    def layout_config(self):
        self.title('Cadastro de Ponto')
        self.geometry('300x300')

    def sistema(self):
        self.title_label = ctk.CTkLabel(self, text='Registro de Ponto', font=('Century Gothic Bold', 24), bg_color='transparent', text_color='#fff')
        self.title_label.place(x=50, y=50)

        self.option = ctk.CTkComboBox(self, values=['Entrada', 'Saída Intervalo', 'Volta Intervalo', 'Saída'], width=195, font=('Century Gothic', 16))
        self.option.set('Entrada')
        self.option.place(x=50, y=145)

        self.nome = ctk.CTkComboBox(self, values=FUNCIONARIOS, width=195, font=('Century Gothic', 16))
        self.nome.set(FUNCIONARIOS[0])
        self.nome.place(x=50, y=100)

        self.bater_ponto = ctk.CTkButton(self, text='Registrar Ponto', font=('Century Gothic', 16), width=195, command=self.submit)
        self.bater_ponto.place(x=50, y=200)

        ficheiro_path = pathlib.Path('Registro de Ponto.xlsx')
        if not ficheiro_path.exists():
            ficheiro = Workbook()
            folha = ficheiro.active
            folha['A1'] = 'Nome Completo'
            folha['B1'] = 'Horário Entrada'
            folha['C1'] = 'Saída Intervalo'
            folha['D1'] = 'Volta Intervalo'
            folha['E1'] = 'Horário Saída'
            folha['F1'] = 'Dia'
            folha['G1'] = 'Horas Trabalhadas'
            folha['H1'] = 'Saldo de Horas'
            self.protect_sheet(folha)
            ficheiro.save('Registro de Ponto.xlsx')

    def protect_sheet(self, sheet):
        # Protege a planilha com uma senha
        sheet.protection.sheet = True
        sheet.protection.password = SENHA

    def submit(self):
        name = self.nome.get()
        option = self.option.get()

        now = datetime.now()
        current_time = now.strftime('%H:%M:%S')
        current_day = now.strftime('%Y-%m-%d')

        ficheiro = openpyxl.load_workbook('Registro de Ponto.xlsx')
        folha = ficheiro.active

        if self.check_existing_entry(folha, name, current_day, option):
            messagebox.showerror("Erro", f"Já existe um registro para {option.lower()} no dia de hoje.")
            return

        found = False
        for row in range(2, folha.max_row + 1):
            if folha.cell(row=row, column=1).value == name and folha.cell(row=row, column=6).value == current_day:
                if option == 'Entrada' and not folha.cell(row=row, column=2).value:
                    folha.cell(row=row, column=2, value=current_time)
                elif option == 'Saída Intervalo' and not folha.cell(row=row, column=3).value:
                    folha.cell(row=row, column=3, value=current_time)
                elif option == 'Volta Intervalo' and not folha.cell(row=row, column=4).value:
                    folha.cell(row=row, column=4, value=current_time)
                elif option == 'Saída' and not folha.cell(row=row, column=5).value:
                    folha.cell(row=row, column=5, value=current_time)
                    self.calculate_worked_hours(folha, row)
                found = True
                break

        if not found:
            new_row = folha.max_row + 1
            folha.cell(column=1, row=new_row, value=name)
            if option == 'Entrada':
                folha.cell(column=2, row=new_row, value=current_time)
            elif option == 'Saída Intervalo':
                folha.cell(column=3, row=new_row, value=current_time)
            elif option == 'Volta Intervalo':
                folha.cell(column=4, row=new_row, value=current_time)
            elif option == 'Saída':
                folha.cell(column=5, row=new_row, value=current_time)
                self.calculate_worked_hours(folha, new_row)
            folha.cell(column=6, row=new_row, value=current_day)

        self.protect_sheet(folha)  # Protege a planilha novamente após a atualização
        ficheiro.save('Registro de Ponto.xlsx')
        messagebox.showinfo("Sistema", 'Dados salvos com sucesso!')
        self.clear()

    def check_existing_entry(self, folha, name, current_day, option):
        for row in range(2, folha.max_row + 1):
            if folha.cell(row=row, column=1).value == name and folha.cell(row=row, column=6).value == current_day:
                if option == 'Entrada' and folha.cell(row=row, column=2).value:
                    return True
                elif option == 'Saída Intervalo' and folha.cell(row=row, column=3).value:
                    return True
                elif option == 'Volta Intervalo' and folha.cell(row=row, column=4).value:
                    return True
                elif option == 'Saída' and folha.cell(row=row, column=5).value:
                    return True
        return False

    def calculate_worked_hours(self, folha, row):
        entrada = folha.cell(row=row, column=2).value
        saida_intervalo = folha.cell(row=row, column=3).value
        volta_intervalo = folha.cell(row=row, column=4).value
        saida = folha.cell(row=row, column=5).value

        if entrada and saida:
            entrada = datetime.strptime(entrada, '%H:%M:%S')
            saida = datetime.strptime(saida, '%H:%M:%S')

            if saida_intervalo and volta_intervalo:
                saida_intervalo = datetime.strptime(saida_intervalo, '%H:%M:%S')
                volta_intervalo = datetime.strptime(volta_intervalo, '%H:%M:%S')
                intervalo = volta_intervalo - saida_intervalo
            else:
                intervalo = timedelta()

            total_trabalhado = saida - entrada - intervalo
            folha.cell(row=row, column=7, value=str(total_trabalhado))

            horas_trabalhadas = total_trabalhado.total_seconds() / 3600
            saldo_horas = round(horas_trabalhadas - 7.20, 2)  
            folha.cell(row=row, column=8, value=f"{saldo_horas:.2f}")

    def clear(self):
        self.nome.set(FUNCIONARIOS[0])
        self.option.set('Entrada')

if __name__ == "__main__":
    app = App()
    app.mainloop()
